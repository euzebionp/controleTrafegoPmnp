from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.shortcuts import render
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import F, Count, Sum, Q
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
from .models import Motorista, Veiculo, Viagem, Multa, Manutencao
from .forms import ViagemForm
from .reports import generate_pdf_report

# Dashboard View
class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'dashboard.html'

# Report Selection View
class ReportSelectionView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/report_selection.html'

# Motorista Views
class MotoristaListView(LoginRequiredMixin, ListView):
    model = Motorista
    template_name = 'drivers/driver_list.html'
    context_object_name = 'motoristas'

class MotoristaCreateView(LoginRequiredMixin, CreateView):
    model = Motorista
    template_name = 'drivers/driver_form.html'
    fields = ['nome', 'cpf', 'cnh', 'validade_cnh']
    success_url = reverse_lazy('motorista_list')

class MotoristaUpdateView(LoginRequiredMixin, UpdateView):
    model = Motorista
    template_name = 'drivers/driver_form.html'
    fields = ['nome', 'cpf', 'cnh', 'validade_cnh']
    success_url = reverse_lazy('motorista_list')

class MotoristaDeleteView(LoginRequiredMixin, DeleteView):
    model = Motorista
    template_name = 'drivers/driver_confirm_delete.html'
    success_url = reverse_lazy('motorista_list')

# Veiculo Views
class VeiculoListView(LoginRequiredMixin, ListView):
    model = Veiculo
    template_name = 'vehicles/vehicle_list.html'
    context_object_name = 'veiculos'

class VeiculoCreateView(LoginRequiredMixin, CreateView):
    model = Veiculo
    template_name = 'vehicles/vehicle_form.html'
    fields = ['placa', 'modelo', 'ano', 'renavam', 'km_atual']
    success_url = reverse_lazy('veiculo_list')

class VeiculoUpdateView(LoginRequiredMixin, UpdateView):
    model = Veiculo
    template_name = 'vehicles/vehicle_form.html'
    fields = ['placa', 'modelo', 'ano', 'renavam', 'km_atual']
    success_url = reverse_lazy('veiculo_list')

class VeiculoDeleteView(LoginRequiredMixin, DeleteView):
    model = Veiculo
    template_name = 'vehicles/vehicle_confirm_delete.html'
    success_url = reverse_lazy('veiculo_list')

# Viagem Views
class ViagemListView(LoginRequiredMixin, ListView):
    model = Viagem
    template_name = 'travels/travel_list.html'
    context_object_name = 'viagens'
    paginate_by = 50
    
    def get_queryset(self):
        return Viagem.objects.select_related(
            'motorista', 'veiculo'
        ).order_by('-data', '-hora_saida')

class ViagemCreateView(LoginRequiredMixin, CreateView):
    model = Viagem
    form_class = ViagemForm
    template_name = 'travels/travel_form.html'
    success_url = reverse_lazy('viagem_list')
    
    def form_valid(self, form):
        form.save(commit=True, request=self.request)
        return super().form_valid(form)

class ViagemUpdateView(LoginRequiredMixin, UpdateView):
    model = Viagem
    form_class = ViagemForm
    template_name = 'travels/travel_form.html'
    success_url = reverse_lazy('viagem_list')
    
    def form_valid(self, form):
        form.save(commit=True, request=self.request)
        return super().form_valid(form)

class ViagemDeleteView(LoginRequiredMixin, DeleteView):
    model = Viagem
    template_name = 'travels/travel_confirm_delete.html'
    success_url = reverse_lazy('viagem_list')

# Multa Views
class MultaListView(LoginRequiredMixin, ListView):
    model = Multa
    template_name = 'fines/fine_list.html'
    context_object_name = 'multas'
    paginate_by = 50
    
    def get_queryset(self):
        return Multa.objects.select_related(
            'motorista', 'veiculo', 'viagem'
        ).order_by('-data')

class MultaCreateView(LoginRequiredMixin, CreateView):
    model = Multa
    template_name = 'fines/fine_form.html'
    fields = ['data', 'hora_infracao', 'local', 'tipo_infracao', 'descricao', 'motorista', 'veiculo', 'viagem', 'valor']
    success_url = reverse_lazy('multa_list')

class MultaUpdateView(LoginRequiredMixin, UpdateView):
    model = Multa
    template_name = 'fines/fine_form.html'
    fields = ['data', 'hora_infracao', 'local', 'tipo_infracao', 'descricao', 'motorista', 'veiculo', 'viagem', 'valor']
    success_url = reverse_lazy('multa_list')

class MultaDeleteView(LoginRequiredMixin, DeleteView):
    model = Multa
    template_name = 'fines/fine_confirm_delete.html'
    success_url = reverse_lazy('multa_list')

# Manutencao Views
class ManutencaoListView(LoginRequiredMixin, ListView):
    model = Manutencao
    template_name = 'maintenance/maintenance_list.html'
    context_object_name = 'manutencoes'
    paginate_by = 50
    
    def get_queryset(self):
        return Manutencao.objects.select_related(
            'veiculo'
        ).order_by('-data')

class ManutencaoCreateView(LoginRequiredMixin, CreateView):
    model = Manutencao
    template_name = 'maintenance/maintenance_form.html'
    fields = ['veiculo', 'data', 'tipo_servico', 'descricao', 'km_realizado', 'proximo_servico_km', 'proximo_servico_data', 'valor']
    success_url = reverse_lazy('manutencao_list')

class ManutencaoUpdateView(LoginRequiredMixin, UpdateView):
    model = Manutencao
    template_name = 'maintenance/maintenance_form.html'
    fields = ['veiculo', 'data', 'tipo_servico', 'descricao', 'km_realizado', 'proximo_servico_km', 'proximo_servico_data', 'valor']
    success_url = reverse_lazy('manutencao_list')

class ManutencaoDeleteView(LoginRequiredMixin, DeleteView):
    model = Manutencao
    template_name = 'maintenance/maintenance_confirm_delete.html'
    success_url = reverse_lazy('manutencao_list')

# Report Export Views
# Report Export Views
def relatorio_motoristas_pdf(request):
    """Exporta relatório de motoristas em PDF"""
    motoristas = Motorista.objects.all().order_by('nome')
    headers = ['Nome', 'CPF', 'CNH', 'Validade CNH']
    data = []
    for m in motoristas:
        data.append([
            m.nome,
            m.cpf,
            m.cnh,
            m.validade_cnh.strftime('%d/%m/%Y') if m.validade_cnh else '-'
        ])
    
    return generate_pdf_report("Relatório de Motoristas", data, headers, "relatorio_motoristas.pdf")

def relatorio_veiculos_pdf(request):
    """Exporta relatório de veículos em PDF"""
    veiculos = Veiculo.objects.all().order_by('placa')
    headers = ['Placa', 'Modelo', 'Ano', 'Renavam', 'KM Atual']
    data = []
    for v in veiculos:
        data.append([
            v.placa,
            v.modelo,
            str(v.ano),
            v.renavam,
            str(v.km_atual)
        ])
    
    return generate_pdf_report("Relatório de Veículos", data, headers, "relatorio_veiculos.pdf")

def relatorio_multas_pdf(request):
    """Exporta relatório de multas em PDF"""
    multas = Multa.objects.select_related('motorista', 'veiculo').order_by('-data')
    headers = ['Data', 'Hora', 'Local', 'Motorista', 'Veículo', 'Valor']
    data = []
    for m in multas:
        data.append([
            m.data.strftime('%d/%m/%Y'),
            m.hora_infracao.strftime('%H:%M') if m.hora_infracao else '-',
            m.local,
            m.motorista.nome if m.motorista else '-',
            m.veiculo.placa if m.veiculo else '-',
            f"R$ {m.valor:.2f}"
        ])
    
    return generate_pdf_report("Relatório de Multas", data, headers, "relatorio_multas.pdf")

def relatorio_manutencoes_pdf(request):
    """Exporta relatório de manutenções em PDF"""
    manutencoes = Manutencao.objects.select_related('veiculo').order_by('-data')
    headers = ['Veículo', 'Data', 'Tipo', 'Descrição', 'Valor']
    data = []
    for m in manutencoes:
        data.append([
            m.veiculo.placa if m.veiculo else '-',
            m.data.strftime('%d/%m/%Y'),
            m.tipo_servico,
            m.descricao,
            f"R$ {m.valor:.2f}"
        ])
    
    return generate_pdf_report("Relatório de Manutenções", data, headers, "relatorio_manutencoes.pdf")

def relatorio_viagens_pdf(request):
    """Exporta relatório de viagens em PDF"""
    viagens = Viagem.objects.select_related('motorista', 'veiculo').order_by('-data')
    headers = ['Data', 'Saída', 'Motorista', 'Veículo', 'Destino']
    data = []
    for v in viagens:
        data.append([
            v.data.strftime('%d/%m/%Y'),
            v.hora_saida.strftime('%H:%M') if v.hora_saida else '-',
            v.motorista.nome if v.motorista else '-',
            v.veiculo.placa if v.veiculo else '-',
            v.destino
        ])
    return generate_pdf_report("Relatório de Viagens", data, headers, "relatorio_viagens.pdf")


# Excel Import/Export Views
import openpyxl
from django.views.generic import FormView, View
from .forms import ViagemImportForm

from openpyxl.worksheet.datavalidation import DataValidation

class DownloadTravelTemplateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modelo Importação Viagens"
        
        # Create hidden sheet for data validation options
        ws_data = wb.create_sheet("Dados")
        ws_data.sheet_state = 'hidden'
        
        # Fetch data
        motoristas = Motorista.objects.all().order_by('nome')
        veiculos = Veiculo.objects.all().order_by('modelo')
        
        driver_options = [f"{m.nome} - {m.cpf}" for m in motoristas]
        vehicle_options = [f"{v.modelo} - {v.placa}" for v in veiculos]
        
        # Write data to hidden sheet (side by side)
        max_rows = max(len(driver_options), len(vehicle_options))
        for i in range(max_rows):
            d_val = driver_options[i] if i < len(driver_options) else ""
            v_val = vehicle_options[i] if i < len(vehicle_options) else ""
            ws_data.append([d_val, v_val])
            
        # Headers
        headers = ['Data (DD/MM/AAAA)', 'Hora Saida (HH:MM)', 'Motorista (Selecione)', 'Veiculo (Selecione)', 'Origem', 'Destino', 'Distancia (KM)', 'KM Final (Atual)']
        ws.append(headers)
        
        # 1. Driver Validation (Column C)
        if driver_options:
            last_row = len(driver_options)
            dv_driver = DataValidation(type="list", formula1=f"'Dados'!$A$1:$A${last_row}", allow_blank=True)
            dv_driver.error = 'Por favor selecione um motorista da lista'
            dv_driver.errorTitle = 'Motorista Inválido'
            dv_driver.add(f'C2:C500')
            ws.add_data_validation(dv_driver)

        # 2. Vehicle Validation (Column D)
        if vehicle_options:
            last_row_v = len(vehicle_options)
            dv_vehicle = DataValidation(type="list", formula1=f"'Dados'!$B$1:$B${last_row_v}", allow_blank=True)
            dv_vehicle.error = 'Por favor selecione um veículo da lista'
            dv_vehicle.errorTitle = 'Veículo Inválido'
            dv_vehicle.add(f'D2:D500')
            ws.add_data_validation(dv_vehicle)
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 25 # Wider for Model - Plate
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=modelo_importacao_viagens.xlsx'
        wb.save(response)
        return response

class ImportTravelView(LoginRequiredMixin, FormView):
    template_name = 'travels/import_travels.html'
    form_class = ViagemImportForm
    success_url = reverse_lazy('viagem_list')
    
    def form_valid(self, form):
        excel_file = form.cleaned_data['arquivo_excel']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            errors = []
            
            # Skip header row
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            for row_idx, row in enumerate(rows, start=2):
                # Check if row is empty
                if not any(row): continue
                
                try:
                    # Unpack expected 8 columns (if 8 exists, else provide default)
                    # Helper to get value securely
                    def get_col(idx):
                        return row[idx] if idx < len(row) else None
                    
                    data_val = get_col(0)
                    hora_val = get_col(1)
                    cpf_val = get_col(2)
                    placa_val = get_col(3)
                    origem_val = get_col(4)
                    destino_val = get_col(5)
                    distancia_val = get_col(6)
                    km_final_val = get_col(7)
                    
                    if not data_val or not hora_val or not cpf_val or not placa_val:
                        errors.append(f"Linha {row_idx}: Campos obrigatórios (Data, Hora, Motorista, Placa) faltando.")
                        continue

                    # Parse 'Nome - CPF' if present
                    raw_cpf = str(cpf_val)
                    if " - " in raw_cpf:
                        raw_cpf = raw_cpf.split(" - ")[-1]

                    # Parse 'Modelo - Placa' if present
                    raw_placa = str(placa_val)
                    if " - " in raw_placa:
                        raw_placa = raw_placa.split(" - ")[-1]
                        
                    # Clean data
                    cpf = raw_cpf.replace('.', '').replace('-', '').strip()
                    # Ensure CPF has 11 digits (pad with zeros if excel stripped them)
                    if len(cpf) < 11:
                         cpf = cpf.zfill(11)

                    # Try searching with formatted CPF as well
                    cpf_formatted = f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
                         
                    placa = str(raw_placa).replace('-', '').strip().upper()
                    # Fallback for concatenated strings without separator (e.g., "MODELO PLACA")
                    # If length > 7, assume the plate is at the end (standard plate is 7 chars)
                    if len(placa) > 7:
                        placa = placa[-7:]
                    
                    # Search valid foreign keys
                    motorista = Motorista.objects.filter(Q(cpf=cpf) | Q(cpf=cpf_formatted)).first()
                    if not motorista:
                        errors.append(f"Linha {row_idx}: Motorista com CPF {cpf} não encontrado.")
                        continue
                        
                    veiculo = Veiculo.objects.filter(placa=placa).first()
                    if not veiculo:
                        errors.append(f"Linha {row_idx}: Veículo com placa {placa} não encontrado.")
                        continue
                    
                    # Handle Date/Time types from Excel
                    # data_val should be datetime or date, hora_val time or datetime
                    import datetime
                    
                    final_date = data_val
                    if isinstance(data_val, str):
                        # Try parsing string date if necessary
                        pass # Let's hope excel gives dates
                    
                    final_time = hora_val
                    if isinstance(hora_val, str):
                        pass
                        
                    # Validate/Clean distance
                    final_distancia = 0
                    if distancia_val:
                        try:
                            if isinstance(distancia_val, str):
                                final_distancia = float(distancia_val.replace(',', '.'))
                            else:
                                final_distancia = float(distancia_val)
                        except (ValueError, TypeError):
                            final_distancia = 0
                            
                    # Validate/Clean KM Final
                    final_km_final = 0
                    if km_final_val:
                        try:
                            if isinstance(km_final_val, str):
                                final_km_final = float(km_final_val.replace(',', '.'))
                            else:
                                final_km_final = float(km_final_val)
                        except (ValueError, TypeError):
                            final_km_final = 0
                        
                    viagem = Viagem.objects.create(
                        data=final_date,
                        hora_saida=final_time,
                        motorista=motorista,
                        veiculo=veiculo,
                        origem=origem_val,
                        destino=destino_val,
                        distancia=final_distancia,
                        km_final=final_km_final
                    )
                    
                    # Update Vehicle KM if provided and greater
                    if final_km_final > float(veiculo.km_atual):
                        veiculo.km_atual = final_km_final
                        veiculo.save()
                        
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"Linha {row_idx}: Erro inesperado - {str(e)}")
            
            if created_count > 0:
                messages.success(self.request, f"{created_count} viagens importadas com sucesso!")
            
            if errors:
                for err in errors[:5]: # Limit errors shown
                    messages.warning(self.request, err)
                if len(errors) > 5:
                    messages.warning(self.request, f"E mais {len(errors)-5} erros. Verifique a planilha.")
                    
        except Exception as e:
            messages.error(self.request, f"Erro ao ler arquivo: {str(e)}")
            return self.form_invalid(form)
            
        return super().form_valid(form)


