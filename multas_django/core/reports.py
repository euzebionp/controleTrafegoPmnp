from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils import timezone

def gerar_relatorio_viagens_pdf(viagens, response):
    """Gera relatório de viagens em PDF"""
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f77b4'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    title = Paragraph("Relatório de Viagens", title_style)
    elements.append(title)
    
    # Data de geração
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    date_text = Paragraph(f"Gerado em: {timezone.now().strftime('%d/%m/%Y %H:%M')}", date_style)
    elements.append(date_text)
    elements.append(Spacer(1, 0.3*inch))
    
    # Dados da tabela
    data = [['Data', 'Hora', 'Motorista', 'Veículo', 'Origem', 'Destino', 'KM']]
    total_km = 0
    
    for v in viagens:
        data.append([
            v.data.strftime('%d/%m/%Y'),
            v.hora_saida.strftime('%H:%M'),
            v.motorista.nome[:25],
            f"{v.veiculo.placa}",
            v.origem[:20],
            v.destino[:20],
            f'{v.distancia:.1f}'
        ])
        total_km += float(v.distancia)
    
    # Linha de total
    data.append(['', '', '', '', '', 'TOTAL:', f'{total_km:.1f} km'])
    
    # Criar tabela
    table = Table(data, colWidths=[1*inch, 0.8*inch, 1.8*inch, 1*inch, 1.5*inch, 1.5*inch, 0.8*inch])
    table.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        # Corpo
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f0f0f0')]),
        # Total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ffeb3b')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('GRID', (0, -1), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    
    # Estatísticas
    elements.append(Spacer(1, 0.3*inch))
    stats_text = Paragraph(
        f"<b>Total de viagens:</b> {len(viagens)} | <b>Quilometragem total:</b> {total_km:.1f} km | <b>Média por viagem:</b> {total_km/len(viagens) if viagens else 0:.1f} km",
        styles['Normal']
    )
    elements.append(stats_text)
    
    doc.build(elements)
    return response

def gerar_relatorio_multas_pdf(multas, response):
    """Gera relatório de multas em PDF"""
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#d32f2f'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    title = Paragraph("Relatório de Multas", title_style)
    elements.append(title)
    
    # Data de geração
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    date_text = Paragraph(f"Gerado em: {timezone.now().strftime('%d/%m/%Y %H:%M')}", date_style)
    elements.append(date_text)
    elements.append(Spacer(1, 0.3*inch))
    
    # Dados da tabela
    data = [['Data', 'Motorista', 'Veículo', 'Tipo de Infração', 'Local', 'Valor']]
    total_valor = 0
    
    for m in multas:
        data.append([
            m.data.strftime('%d/%m/%Y'),
            m.motorista.nome[:25],
            m.veiculo.placa,
            m.tipo_infracao[:30],
            m.local[:25],
            f'R$ {m.valor:.2f}'
        ])
        total_valor += float(m.valor)
    
    # Linha de total
    data.append(['', '', '', '', 'TOTAL:', f'R$ {total_valor:.2f}'])
    
    # Criar tabela
    table = Table(data, colWidths=[1*inch, 1.8*inch, 1*inch, 2*inch, 1.8*inch, 1*inch])
    table.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d32f2f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        # Corpo
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#ffebee')]),
        # Total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ffeb3b')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('GRID', (0, -1), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    
    # Estatísticas
    elements.append(Spacer(1, 0.3*inch))
    stats_text = Paragraph(
        f"<b>Total de multas:</b> {len(multas)} | <b>Valor total:</b> R$ {total_valor:.2f} | <b>Valor médio:</b> R$ {total_valor/len(multas) if multas else 0:.2f}",
        styles['Normal']
    )
    elements.append(stats_text)
    
    doc.build(elements)
    return response

def gerar_relatorio_viagens_excel(viagens, response):
    """Gera relatório de viagens em Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Viagens"
    
    # Cabeçalhos
    headers = ['Data', 'Hora Saída', 'Motorista', 'Veículo', 'Origem', 'Destino', 'Distância (km)']
    ws.append(headers)
    
    # Estilizar cabeçalho
    header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    total_km = 0
    for v in viagens:
        ws.append([
            v.data.strftime('%d/%m/%Y'),
            v.hora_saida.strftime('%H:%M'),
            v.motorista.nome,
            f"{v.veiculo.placa} - {v.veiculo.modelo}",
            v.origem,
            v.destino,
            float(v.distancia)
        ])
        total_km += float(v.distancia)
    
    # Linha de total
    last_row = ws.max_row + 1
    ws.append(['', '', '', '', '', 'TOTAL:', total_km])
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="ffeb3b", end_color="ffeb3b", fill_type="solid")
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15
    
    wb.save(response)
    return response

def gerar_relatorio_multas_excel(multas, response):
    """Gera relatório de multas em Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Multas"
    
    # Cabeçalhos
    headers = ['Data', 'Motorista', 'Veículo', 'Tipo de Infração', 'Local', 'Valor (R$)']
    ws.append(headers)
    
    # Estilizar cabeçalho
    header_fill = PatternFill(start_color="d32f2f", end_color="d32f2f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    total_valor = 0
    for m in multas:
        ws.append([
            m.data.strftime('%d/%m/%Y'),
            m.motorista.nome,
            f"{m.veiculo.placa} - {m.veiculo.modelo}",
            m.tipo_infracao,
            m.local,
            float(m.valor)
        ])
        total_valor += float(m.valor)
    
    # Linha de total
    last_row = ws.max_row + 1
    ws.append(['', '', '', '', 'TOTAL:', total_valor])
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="ffeb3b", end_color="ffeb3b", fill_type="solid")
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    
    wb.save(response)
    return response
